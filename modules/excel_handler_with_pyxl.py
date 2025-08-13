"""
Excel handler module - openpyxl based implementation
Step by step implementation for format preservation and data accuracy
"""

import openpyxl
import msoffcrypto
import io
import os
import zipfile
import tempfile
from datetime import datetime, timedelta
from typing import Optional
import pandas as pd

EXCEL_EPOCH = datetime(1899, 12, 30)  # 1900 시스템
LEAP_BUG_CUTOFF = datetime(1900, 3, 1)
DEFAULT_EXCEL_PATH = "./매출리포트-250810203219_1 - Sample.xlsx"

def excel_serial_to_datetime(serial: float) -> datetime:
    """엑셀 시리얼(정수+분수)을 datetime으로 변환"""
    return EXCEL_EPOCH + timedelta(days=float(serial))

def excel_serial_to_str(serial: float, with_time: bool = True) -> str:
    dt = excel_serial_to_datetime(serial)
    return dt.strftime("%Y-%m-%d %H:%M:%S" if with_time else "%Y-%m-%d")

def _dt_to_excel_serial(dt: datetime) -> int | float:
    serial = (dt - EXCEL_EPOCH).days + (dt - dt.replace(hour=0, minute=0, second=0, microsecond=0)).total_seconds()/86400.0
    # 1900-01-01 ~ 1900-02-28 구간 보정
    if datetime(1900, 1, 1) <= dt < LEAP_BUG_CUTOFF:
        serial -= 1
    # 정수면 int로
    return int(serial) if abs(serial - round(serial)) < 1e-9 else serial

class ExcelHandlerPyXL:
    def __init__(self, excel_path: str, password: str = None):
        """Initialize Excel handler"""
        self.excel_path = excel_path
        self.password = password
        self.workbook: Optional[openpyxl.Workbook] = None
        self.worksheet: Optional[openpyxl.worksheet.worksheet.Worksheet] = None
        
        # Check file existence
        if not os.path.exists(excel_path):
            raise FileNotFoundError(f"File not found: {excel_path}")
        
        print(f"[FILE] Excel file: {excel_path}")
        print(f"[AUTH] Password: {'SET' if password else 'NONE'}")

    # -------------------------------
    # 내부 유틸: 파일 열기 & styles.xml 제거 & 편집 보장
    # -------------------------------

    def _decrypt_if_needed(self) -> str:
        """
        암호가 있으면 복호화 후 임시 파일 경로를 반환,
        없으면 원본 경로를 반환.
        """
        if not self.password:
            return self.excel_path

        print("[DECRYPT] Password-protected file detected. Decrypting to temp file...")
        with open(self.excel_path, "rb") as f:
            office = msoffcrypto.OfficeFile(f)
            office.load_key(password=self.password)
            out = io.BytesIO()
            try:
                office.decrypt(out)
                print("[OK] decrypt() method successful")
            except Exception as e:
                print(f"[WARN] decrypt() failed: {e}")
                print("[RETRY] Trying save() method...")
                out = io.BytesIO()
                office.save(out)
                print("[OK] save() method successful")
            out.seek(0)

        # Bytes → 임시 파일로 저장 (openpyxl은 파일경로가 다루기 편함)
        tmp_path = os.path.join(tempfile.mkdtemp(prefix="xlsx_dec_"), "decrypted.xlsx")
        with open(tmp_path, "wb") as fw:
            fw.write(out.getbuffer())
        return tmp_path

    def _remove_styles_xml_copy(self, src_path: str) -> str:
        """
        xlsx(zip)에서 xl/styles.xml만 제거한 임시 사본을 만들어 경로 반환.
        """
        tmp_dir = tempfile.mkdtemp(prefix="xlsx_repair_")
        repaired_path = os.path.join(tmp_dir, os.path.basename(src_path))
        with zipfile.ZipFile(src_path, "r") as zin, \
             zipfile.ZipFile(repaired_path, "w", compression=zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                if item.filename == "xl/styles.xml":
                    # styles.xml 제거
                    continue
                zout.writestr(item, zin.read(item.filename))
        return repaired_path

    def _clone_readonly_to_editable(self, ro_path: str) -> openpyxl.Workbook:
        """
        read_only=True로만 열리는 파일을 '데이터만' 새 워크북으로 복사해 편집 가능하게 만든다.
        내부 값 우선(internal_value)로 복사하되, EmptyCell 등은 value→None 순으로 안전 처리.
        """
        print("[FALLBACK] Cloning read-only workbook to an editable workbook (values only, using internal_value safely)...")
        ro_wb = openpyxl.load_workbook(ro_path, read_only=True, data_only=False)
        new_wb = openpyxl.Workbook()
        new_wb.remove(new_wb.active)

        def _raw(c):
            # internal_value가 있으면 그걸 우선 사용
            if hasattr(c, "internal_value"):
                return c.internal_value
            # 없으면 일반 value (ReadOnlyCell/EmptyCell 호환)
            return getattr(c, "value", None)

        for name in ro_wb.sheetnames:
            src_ws = ro_wb[name]
            dst_ws = new_wb.create_sheet(title=name)
            for row in src_ws.iter_rows(values_only=False):
                dst_ws.append([_raw(c) for c in row])

        return new_wb

    def _load_editable_workbook(self, base_path: str) -> openpyxl.Workbook:
        """
        1) read_only=False 직접 로드
        2) styles.xml 제거 사본으로 read_only=False 재시도
        3) 실패 시 read_only=True로 열어 값만 새 워크북으로 복제
        """
        e1 = e2 = e3 = None  # 예외 변수 사전 초기화

        # 1) 직접 시도
        try:
            print("[ATTEMPT] Direct load (read_only=False)")
            wb = openpyxl.load_workbook(base_path, read_only=False, data_only=False, keep_links=True)
            print("[OK] Direct load successful")
            return wb
        except Exception as err:
            print(f"[ERROR 1] Direct load failed:")
            print(f"  - Error type: {type(err).__name__}")
            print(f"  - Error message: {str(err)}")
            print(f"  - File path: {base_path}")
            print(f"  - File exists: {os.path.exists(base_path)}")
            if os.path.exists(base_path):
                print(f"  - File size: {os.path.getsize(base_path)} bytes")
            
            # 더 상세한 traceback 출력
            import traceback
            print(f"  - Full traceback:")
            traceback.print_exc()

        # 2) styles.xml 제거 사본으로 재시도
        try:
            print("[REPAIR] Making styles-stripped copy and retrying read_only=False...")
            repaired = self._remove_styles_xml_copy(base_path)
            wb = openpyxl.load_workbook(repaired, read_only=False, data_only=False, keep_links=True)
            print("[OK] Repaired copy load successful")
            return wb
        except Exception as err:
            e2 = err
            print(f"[WARN] Repaired copy load failed: {type(e2).__name__}: {e2}")

        # 3) 최후: read_only=True → 값만 복제
        try:
            wb = self._clone_readonly_to_editable(base_path)
            print("[OK] Cloned to an editable workbook (values only)")
            return wb
        except Exception as err:
            e3 = err

        # 실패 보고(정의된 예외만 안전하게 포함)
        parts = []
        if e1 is not None: parts.append(f"Direct: {type(e1).__name__}: {e1}")
        if e2 is not None: parts.append(f"Repaired: {type(e2).__name__}: {e2}")
        if e3 is not None: parts.append(f"Clone: {type(e3).__name__}: {e3}")
        summary = " | ".join(parts) if parts else "Unknown error"

        raise RuntimeError(f"Failed to obtain an editable workbook. {summary}")
        
    # -------------------------------
    # 공개 API
    # -------------------------------

    def read_excel_basic(self):
        """Basic Excel file reading with password support + editable guarantee"""
        try:
            print("[STEP1] Preparing source...")
            src_path = self._decrypt_if_needed()  # 암호 있으면 복호화 임시 파일, 없으면 원본 경로

            print("[STEP2] Loading workbook with editable guarantee...")
            self.workbook = self._load_editable_workbook(src_path)

            # Show available sheets
            print(f"[SHEETS] Available sheets: {self.workbook.sheetnames}")
            
            # Select worksheet
            target_sheet = "상품 주문 상세내역"
            if target_sheet in self.workbook.sheetnames:
                self.worksheet = self.workbook[target_sheet]
                print(f"[SELECT] Using target sheet: '{target_sheet}'")
            else:
                first_sheet = self.workbook.sheetnames[0]
                self.worksheet = self.workbook[first_sheet]
                print(f"[SELECT] Using first sheet: '{first_sheet}'")
            
            print(f"[INFO] Max row: {self.worksheet.max_row}")
            print(f"[INFO] Max column: {self.worksheet.max_column}")
            
            # 편집 가능 여부 점검
            try:
                cell = self.worksheet.cell(1, 1)
                _old = cell.value
                self.worksheet.cell(1, 1, value=_old)  # no-op set
                print("[INFO] Worksheet is EDITABLE (read_only=False)")
            except Exception:
                print("[INFO] Worksheet seems READ-ONLY")

            return True
            
        except Exception as e:
            print(f"[ERROR] Excel reading failed: {e}")
            return False
        

    def switch_to_sheet(self, sheet_name: str) -> bool:
        """
        워킹 시트를 지정된 시트로 변경
        @param sheet_name: 변경할 시트명
        @returns: 변경 성공 여부
        """
        if not self.workbook:
            print(f"[ERROR] 워크북이 로드되지 않았습니다.")
            return False
            
        if sheet_name not in self.workbook.sheetnames:
            print(f"[ERROR] 시트 '{sheet_name}'를 찾을 수 없습니다.")
            print(f"[INFO] 사용 가능한 시트: {self.workbook.sheetnames}")
            return False
        
        self.worksheet = self.workbook[sheet_name]
        print(f"[INFO] 워킹 시트를 '{sheet_name}'로 변경했습니다.")
        return True
        

    def add_delivery_columns_to_df(self, df: pd.DataFrame) -> pd.DataFrame:
        """배송 정보 컬럼 추가"""
        delivery_cols = [
            '수하인명', '수하인주소', '수하인전화번호', '수하인핸드폰번호',
            '박스수량', '택배운임', '운임구분', '품목명', '배송메세지'
        ]
        for col in delivery_cols:
            if col not in df.columns:
                df[col] = None
        return df

    def filter_to_new_sheet_raw(self,
                            keywords: list[str],
                            new_sheet_name: str = "필터링_결과",
                            mode: str = "any",
                            extra_cols: dict[str, object] | None = None,
                            save_path: str | None = None) -> str:
        """
        - 현재 선택된 worksheet를 raw로 읽어 DataFrame 구성(.internal_value 사용)
        - 옵션 컬럼에서 keywords로 필터(any/all)
        - '수량'은 숫자 그대로, 날짜 컬럼은 시리얼→문자열로 변환
        - extra_cols 신규 컬럼 추가
        - 새 시트에 기록 후 저장
        """
        if self.workbook is None or self.worksheet is None:
            raise RuntimeError("워크북/워크시트가 로드되지 않았습니다. read_excel_basic()을 먼저 호출하세요.")

        # 1) raw DataFrame
        df = self._sheet_to_dataframe_raw(self.worksheet)
        if df.empty:
            raise ValueError("원본 시트가 비어 있습니다.")

        # 2) 옵션 컬럼
        option_col = self._find_option_colname(df)
        if option_col is None:
            raise KeyError("헤더에 '옵션'이 포함된 컬럼을 찾지 못했습니다.")

        # 3) 필터
        s = df[option_col].fillna("").astype(str)
        if mode == "all":
            mask = True
            for kw in keywords:
                mask = mask & s.str.contains(kw, na=False)
        else:
            mask = False
            for kw in keywords:
                mask = mask | s.str.contains(kw, na=False)

        filtered = df[mask].copy()

        # 4) 타입 정리: 수량/날짜
        # 수량: 숫자 그대로(문자면 숫자로 캐스팅 시도)
        if "수량" in filtered.columns:
            def _fix_qty(v):
                if v is None or v == "":
                    return ""
                # 1) 이미 숫자
                if isinstance(v, (int, float)):
                    return int(v) if float(v).is_integer() else float(v)
                # 2) datetime → 시리얼(=원래 숫자)
                if isinstance(v, datetime):
                    return _dt_to_excel_serial(v)
                # 3) 'YYYY-MM-DD...' 문자열 → datetime 파싱 후 시리얼
                if isinstance(v, str) and len(v) >= 10 and v[4] == "-" and v[7] == "-":
                    try:
                        # 시간 포함/미포함 모두 처리
                        fmt = "%Y-%m-%d %H:%M:%S" if " " in v else "%Y-%m-%d"
                        dt = datetime.strptime(v[:19], fmt) if " " in v else datetime.strptime(v[:10], fmt)
                        return _dt_to_excel_serial(dt)
                    except Exception:
                        pass
                # 4) 숫자문자
                try:
                    f = float(str(v).replace(",", ""))
                    return int(f) if f.is_integer() else f
                except Exception:
                    return v
            filtered["수량"] = filtered["수량"].map(_fix_qty)

        # 날짜 컬럼들: 시리얼 → 문자열(시/분/초 필요 여부 선택)
        # date_cols = []
        # for name in ["주문기준일자", "주문시작시각"]:
        #     if name in filtered.columns:
        #         date_cols.append(name)

        # for col in date_cols:
        #     def _fix_date(v):
        #         if v is None or v == "":
        #             return ""
        #         # raw 시리얼이면 숫자/문자 숫자로 처리
        #         try:
        #             f = float(v)
        #             # 주문시작시각은 보통 시간 포함 → with_time=True
        #             with_time = (col == "주문시작시각")
        #             return excel_serial_to_str(f, with_time=with_time)
        #         except Exception:
        #             # 이미 사람이 읽을 수 있는 값이면 그대로
        #             return v
        #     filtered[col] = filtered[col].map(_fix_date)

        # 5) 추가 컬럼
        if extra_cols:
            for c, val in extra_cols.items():
                filtered[c] = val

        # 6) 기존 동명 시트 있으면 삭제(선택)
        if new_sheet_name in self.workbook.sheetnames:
            del self.workbook[new_sheet_name]

        filtered = self.add_delivery_columns_to_df(filtered)

        # 7) 새 시트에 기록
        ws_new = self.workbook.create_sheet(title=new_sheet_name)
        ws_new.append(list(filtered.columns))
        for row in filtered.itertuples(index=False, name=None):
            ws_new.append(row)


        # 8) 저장
        if save_path is None:
            base, ext = os.path.splitext(self.excel_path)
            save_path = f"{base}_filtered.xlsx"
        self.workbook.save(save_path)
        print(f"[SAVE] '{new_sheet_name}' 저장 완료: {save_path} (rows={len(filtered)})")
        return save_path

    # --------------------------
    # 내부 메서드
    # --------------------------
    def _sheet_to_dataframe_raw(self, ws) -> pd.DataFrame:
        """values_only=False로 Cell을 받아 raw/internal_value 우선으로 DataFrame 구성 (EmptyCell 안전)"""
        rows = list(ws.iter_rows(values_only=False))
        if not rows:
            return pd.DataFrame()

        def _raw(c):
            if hasattr(c, "internal_value"):
                return c.internal_value
            return getattr(c, "value", None)

        # 1행 = 헤더
        header = [str(_raw(c)) if _raw(c) is not None else "" for c in rows[0]]

        data = []
        for row in rows[1:]:
            data.append([_raw(c) for c in row])

        return pd.DataFrame(data, columns=header)

    def _find_option_colname(self, df: pd.DataFrame) -> str | None:
        for col in df.columns:
            if "옵션" in str(col):
                return col
        return None

# Test functions
def test_init():
    """Test initialization"""
    try:
        handler = ExcelHandlerPyXL(DEFAULT_EXCEL_PATH, "1202")
        print("[OK] Initialization successful")
        return handler
    except Exception as e:
        print("[ERROR] Initialization failed: {e}")
        return None

def test_read():
    """Test Excel reading"""
    print("=" * 50)
    print("STEP 2: Excel Reading Test")
    print("=" * 50)
    
    handler = ExcelHandlerPyXL(DEFAULT_EXCEL_PATH, None)  # No password
    
    if handler.read_excel_basic():
        print("[SUCCESS] Excel file read successfully")
        return handler
    else:
        print("[FAILED] Excel file reading failed")
        return None

def test_no_password():
    """Test without password - should work much better"""
    print("=" * 60)
    print("STEP 2-B: No Password Excel Reading Test")
    print("=" * 60)
    
    handler = ExcelHandlerPyXL(DEFAULT_EXCEL_PATH, None)

    if handler.read_excel_basic():
        handler.filter_to_new_sheet_raw(
            keywords=["채널추가무료배송", "택배요청"],
            new_sheet_name="필터링_결과",
            mode="any",  # 'all'이면 두 키워드 모두 포함한 행만
            extra_cols={"배송처리상태": "대기", "메모": ""},  # 새 열 추가
            save_path="./매출리포트-250810203219_1_filtered.xlsx"
        )

def test_data_inspection():
    """엑셀 데이터 형태 확인을 위한 테스트"""
    print("=" * 60)
    print("STEP: 데이터 형태 확인 테스트")
    print("=" * 60)
    
    handler = ExcelHandlerPyXL(DEFAULT_EXCEL_PATH, None)

    if handler.read_excel_basic():
        print("\n[1] 워크시트 기본 정보:")
        print(f"   - 시트명: {handler.worksheet.title}")
        print(f"   - 최대 행: {handler.worksheet.max_row}")
        print(f"   - 최대 열: {handler.worksheet.max_column}")
        
        # raw DataFrame으로 읽기
        df = handler._sheet_to_dataframe_raw(handler.worksheet)
        
        print(f"\n[2] DataFrame 기본 정보:")
        print(f"   - 행 수: {len(df)}")
        print(f"   - 열 수: {len(df.columns)}")
        
        print(f"\n[3] 컬럼명 확인:")
        for i, col in enumerate(df.columns):
            print(f"   {i:2d}: '{col}'")
        
        # 핵심 컬럼들 찾기
        target_columns = ['주문기준일자', '주문시작시각', '상품명', '옵션']
        found_columns = {}
        
        print(f"\n[4] 핵심 컬럼 데이터 형태 확인:")
        for target in target_columns:
            # 부분 매칭으로 컬럼 찾기
            found_col = None
            for col in df.columns:
                if target in str(col):
                    found_col = col
                    break
            
            if found_col:
                found_columns[target] = found_col
                print(f"\n   ✅ '{target}' 컬럼 발견: '{found_col}'")
                
                # 첫 5개 행의 데이터 확인
                sample_data = df[found_col].head(5)
                for idx, value in enumerate(sample_data):
                    print(f"      {idx+1}: {repr(value)} (type: {type(value).__name__})")
            else:
                print(f"\n   ❌ '{target}' 컬럼을 찾지 못함")
        
        # 옵션 컬럼에서 택배요청/채널추가무료배송 포함 데이터 확인
        if '옵션' in found_columns:
            option_col = found_columns['옵션']
            print(f"\n[5] 옵션 컬럼 특정 데이터 확인:")
            
            # 택배요청 또는 채널추가무료배송 포함하는 행 찾기
            mask = df[option_col].fillna("").astype(str).str.contains("택배요청|채널추가무료배송", na=False)
            filtered_data = df[mask]
            
            print(f"   - 택배/채널 관련 행 수: {len(filtered_data)}")
            
            if len(filtered_data) > 0:
                print(f"   - 첫 3개 행의 핵심 데이터:")
                for idx, (_, row) in enumerate(filtered_data.head(3).iterrows()):
                    print(f"\n     행 {idx+1}:")
                    for target, col_name in found_columns.items():
                        value = row[col_name]
                        print(f"       {target}: {repr(value)} ({type(value).__name__})")
        
        print(f"\n[6] 영수증 데이터와 비교:")
        receipt_example = {
            "approved_at": "2025-08-01 11:14:31",
            "items": [{"name": "주이패턴이불(냉감나일론)"}],
            "options": "택배요청(0)/민트(0)/Q(20000)"
        }
        print(f"   영수증 시간: {receipt_example['approved_at']}")
        print(f"   영수증 상품: {receipt_example['items'][0]['name']}")
        print(f"   영수증 옵션: {receipt_example['options']}")
        
        return handler, df, found_columns
    else:
        print("[FAILED] Excel 파일 읽기 실패")
        return None, None, None

# if __name__ == "__main__":
    # test_init()
    # test_read()
    # test_no_password()
    # test_data_inspection()